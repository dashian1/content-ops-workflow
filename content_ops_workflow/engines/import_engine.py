from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import threading
import time
import urllib.request
from urllib.parse import quote

from content_ops_workflow.config import SETTINGS


LOGIN_WAIT_SECONDS = 600
COLLECT_WAIT_SECONDS = 45
URL_RE = re.compile(r"https?://[^\s<>'\"，。；、）)\]]+")


def _job_paths(job_id: str) -> tuple[str, str]:
    return (
        os.path.join(SETTINGS.output_dir, "_jobs", f"import_stat_{job_id}.txt"),
        os.path.join(SETTINGS.output_dir, "_jobs", f"import_res_{job_id}.json"),
    )


def _write_status(path: str, status: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(status)


def _safe_filename(value: str, fallback: str = "item") -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    return cleaned[:70].strip(" ._") or fallback


def _unique_dir(parent: str, name: str) -> str:
    path = os.path.join(parent, name)
    if not os.path.exists(path):
        return path
    for index in range(2, 1000):
        candidate = os.path.join(parent, f"{name}_{index}")
        if not os.path.exists(candidate):
            return candidate
    return os.path.join(parent, f"{name}_{int(time.time())}")


def detect_platform(url: str) -> str:
    lowered = url.lower()
    if "douyin.com" in lowered:
        return "抖音"
    if "xiaohongshu.com" in lowered or "xhslink.com" in lowered:
        return "小红书"
    if "bilibili.com" in lowered or "b23.tv" in lowered:
        return "B站"
    if "tiktok.com" in lowered:
        return "TikTok"
    if "kuaishou.com" in lowered:
        return "快手"
    if "youtube.com" in lowered or "youtu.be" in lowered:
        return "YouTube"
    return "未知平台"


def _extract_external_id(url: str) -> str:
    patterns = [
        r"douyin\.com/video/(\d+)",
        r"xiaohongshu\.com/(?:explore|discovery/item)/([^/?#]+)",
        r"bilibili\.com/video/([^/?#]+)",
        r"tiktok\.com/@[^/]+/video/(\d+)",
        r"youtu\.be/([^/?#]+)",
        r"youtube\.com/watch\?v=([^&#]+)",
    ]
    for pattern in patterns:
        matched = re.search(pattern, url, re.I)
        if matched:
            return matched.group(1)
    return hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]


def _make_manual_record(url: str, desc: str, source: str, index: int) -> dict[str, object]:
    platform = detect_platform(url)
    external_id = _extract_external_id(url)
    title = desc.strip() or f"{platform}素材 {index}"
    return {
        "aweme_id": external_id,
        "external_id": external_id,
        "platform": platform,
        "source": source,
        "desc": title,
        "title": title,
        "create_time": 0,
        "duration": 0,
        "author_name": "",
        "author_id": "",
        "like_count": 0,
        "comment_count": 0,
        "collect_count": 0,
        "share_count": 0,
        "video_url": url,
        "cover_url": "",
        "page_url": url,
        "keyword": "",
    }


def _records_from_text(raw: str, source: str) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    seen: set[str] = set()
    for line in raw.splitlines():
        urls = URL_RE.findall(line)
        if not urls:
            continue
        desc = URL_RE.sub("", line).strip(" -｜|:：\t")
        for url in urls:
            clean_url = url.strip()
            if clean_url in seen:
                continue
            seen.add(clean_url)
            records.append(_make_manual_record(clean_url, desc, source, len(records) + 1))
    if not records:
        for url in URL_RE.findall(raw):
            clean_url = url.strip()
            if clean_url in seen:
                continue
            seen.add(clean_url)
            records.append(_make_manual_record(clean_url, "", source, len(records) + 1))
    return records


def _write_records(records: list[dict[str, object]], source: str) -> dict[str, object]:
    today = time.strftime("%Y-%m-%d")
    out_dir = os.path.join(SETTINGS.output_dir, "素材导入", today, f"{source}_{int(time.time())}")
    os.makedirs(out_dir, exist_ok=True)
    archived: list[dict[str, object]] = []
    for index, record in enumerate(records, 1):
        desc = str(record.get("desc") or record.get("title") or "")
        platform = str(record.get("platform", "素材"))
        folder_name = f"{index:02d}_{_safe_filename(platform + '_' + desc, str(index))}"
        folder = _unique_dir(out_dir, folder_name)
        os.makedirs(folder, exist_ok=True)
        record = dict(record)
        record["archive_dir"] = folder
        with open(os.path.join(folder, "metadata.json"), "w", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False, indent=2)
        archived.append(record)
    return {"ok": True, "count": len(archived), "videos": archived, "out_dir": out_dir}


def ingest_browser_text(raw: str) -> dict[str, object]:
    records = _records_from_text(raw.strip(), "browser")
    if not records:
        return {"ok": False, "error": "没有识别到链接。请在已登录的平台页面复制视频链接、分享链接或包含链接的页面内容。"}
    return _write_records(records, "browser")


def ingest_links(raw: str) -> dict[str, object]:
    records = _records_from_text(raw.strip(), "links")
    if not records:
        return {"ok": False, "error": "没有识别到链接。请一行一个粘贴抖音、小红书、B站、TikTok 等素材链接。"}
    return _write_records(records, "links")


def _first_value(row: dict[str, object], names: tuple[str, ...]) -> str:
    normalized = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        value = normalized.get(name.lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _int_value(row: dict[str, object], names: tuple[str, ...]) -> int:
    raw = _first_value(row, names).replace(",", "")
    try:
        return int(float(raw)) if raw else 0
    except ValueError:
        return 0


def _record_from_table_row(row: dict[str, object], index: int) -> dict[str, object] | None:
    url = _first_value(row, ("url", "link", "链接", "素材链接", "视频链接", "页面链接"))
    if not url:
        return None
    desc = _first_value(row, ("desc", "title", "标题", "文案", "内容", "视频标题"))
    record = _make_manual_record(url, desc, "table", index)
    record["author_name"] = _first_value(row, ("author", "作者", "账号", "达人", "nickname"))
    record["like_count"] = _int_value(row, ("like", "likes", "点赞", "点赞数"))
    record["comment_count"] = _int_value(row, ("comment", "comments", "评论", "评论数"))
    record["collect_count"] = _int_value(row, ("collect", "收藏", "收藏数"))
    record["share_count"] = _int_value(row, ("share", "分享", "分享数"))
    record["sales"] = _first_value(row, ("sales", "销量", "成交", "GMV", "询单"))
    record["keyword"] = _first_value(row, ("keyword", "关键词", "标签"))
    platform = _first_value(row, ("platform", "平台"))
    if platform:
        record["platform"] = platform
    return record


def _read_csv_rows(path: str) -> list[dict[str, object]]:
    for encoding in ("utf-8-sig", "gbk"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    with open(path, newline="", encoding="utf-8", errors="ignore") as f:
        return list(csv.DictReader(f))


def _read_xlsx_rows(path: str) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("当前环境没有 openpyxl，Excel 请先另存为 CSV 后导入。") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip() for x in rows[0]]
    return [{headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))} for values in rows[1:]]


def import_table(table_path: str) -> dict[str, object]:
    path = table_path.strip().strip('"')
    if not path or not os.path.exists(path):
        return {"ok": False, "error": "表格路径不存在。请填写本机 CSV 或 Excel 文件路径。"}
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        rows = _read_csv_rows(path)
    elif ext in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(path)
    else:
        return {"ok": False, "error": "当前支持 CSV、XLSX、XLSM。Excel 导入失败时可先另存为 CSV。"}
    records = [record for i, row in enumerate(rows, 1) if (record := _record_from_table_row(row, i))]
    if not records:
        return {"ok": False, "error": "表格里没有识别到链接列。列名可用：链接、素材链接、视频链接、url、link。"}
    return _write_records(records, "table")


def _extract_videos(api_responses: list[dict[str, object]], keyword: str, max_count: int) -> list[dict[str, object]]:
    videos: list[dict[str, object]] = []
    seen: set[str] = set()
    for resp in api_responses:
        items = resp.get("data", [])
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            aweme = item.get("aweme_info") or item
            if not isinstance(aweme, dict):
                continue
            aid = str(aweme.get("aweme_id", ""))
            if not aid or aid in seen:
                continue
            seen.add(aid)
            author = aweme.get("author", {}) or {}
            stats = aweme.get("statistics", {}) or {}
            video = aweme.get("video", {}) or {}
            desc = aweme.get("desc", "")
            videos.append(
                {
                    "aweme_id": aid,
                    "external_id": aid,
                    "platform": "抖音",
                    "source": "douyin_search",
                    "desc": desc,
                    "title": desc,
                    "create_time": aweme.get("create_time", 0),
                    "duration": video.get("duration", 0) // 1000 if video.get("duration") else 0,
                    "author_name": author.get("nickname", ""),
                    "author_id": str(author.get("uid", "")),
                    "like_count": stats.get("digg_count", 0),
                    "comment_count": stats.get("comment_count", 0),
                    "collect_count": stats.get("collect_count", 0),
                    "share_count": stats.get("share_count", 0),
                    "video_url": ((video.get("download_addr", {}) or {}).get("url_list") or (video.get("play_addr", {}) or {}).get("url_list") or [""])[0],
                    "cover_url": ((video.get("cover", {}) or {}).get("url_list") or [""])[0],
                    "page_url": f"https://www.douyin.com/video/{aid}",
                    "keyword": keyword,
                }
            )
            if len(videos) >= max_count:
                return videos
    return videos


def _safe_goto(page, url: str) -> None:
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
    except Exception:
        page.wait_for_timeout(2000)


def _safe_scroll(page) -> bool:
    if page.is_closed():
        raise RuntimeError("抖音窗口被关闭，采集已停止")
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        return True
    except Exception as exc:
        message = str(exc)
        if "Execution context was destroyed" in message or "navigation" in message.lower():
            page.wait_for_timeout(2500)
            return False
        raise


def _safe_body_text(page) -> str:
    if page.is_closed():
        raise RuntimeError("抖音窗口被关闭，采集已停止")
    try:
        return page.locator("body").inner_text(timeout=5000)[:2000]
    except Exception as exc:
        message = str(exc)
        if "Execution context was destroyed" in message or "navigation" in message.lower():
            page.wait_for_timeout(2500)
        return ""


def _download_if_possible(url: object, path: str, timeout: int) -> None:
    if not url:
        return
    try:
        req = urllib.request.Request(str(url), headers={"Referer": "https://www.douyin.com/", "User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            with open(path, "wb") as f:
                shutil.copyfileobj(resp, f)
    except Exception:
        pass


def start_douyin_search(keyword: str, max_count: int = 10) -> dict[str, object]:
    keyword = keyword.strip()
    if not keyword:
        return {"ok": False, "error": "请输入关键词"}
    job_id = str(int(time.time()))

    def run() -> None:
        stat_path, res_path = _job_paths(job_id)
        _write_status(stat_path, "starting:正在打开抖音浏览器")
        safe_kw = re.sub(r'[\\/:*?"<>|]', "_", keyword)[:30]
        try:
            from playwright.sync_api import sync_playwright
            from playwright_stealth import Stealth

            user_data = os.path.join(SETTINGS.root, ".runtime", "douyin_profile")
            os.makedirs(user_data, exist_ok=True)
            api_responses: list[dict[str, object]] = []

            with sync_playwright() as p:
                ctx = p.chromium.launch_persistent_context(user_data_dir=user_data, headless=False, viewport={"width": 1440, "height": 900})
                page = ctx.new_page()
                Stealth().apply_stealth_sync(page)

                def on_resp(resp) -> None:
                    if "/aweme/v1/web/search/item/" in resp.url and resp.status == 200:
                        try:
                            api_responses.append(resp.json())
                        except Exception:
                            pass

                page.on("response", on_resp)
                search_url = f"https://www.douyin.com/search/{quote(keyword)}?type=video"
                _safe_goto(page, search_url)

                videos: list[dict[str, object]] = []
                deadline = time.time() + LOGIN_WAIT_SECONDS
                while time.time() < deadline and len(videos) < max_count:
                    _write_status(stat_path, "collecting:正在采集搜索结果")
                    for _ in range(3):
                        _safe_scroll(page)
                        time.sleep(2)
                    videos = _extract_videos(api_responses, keyword, max_count)
                    if videos:
                        break
                    body_text = _safe_body_text(page)
                    if "登录" in body_text or "验证码" in body_text or "扫码" in body_text:
                        _write_status(stat_path, "login_required:请在打开的抖音窗口完成登录，登录后系统会自动继续采集")
                    else:
                        _write_status(stat_path, "waiting:还没有拿到搜索数据，请保持抖音窗口打开")
                    time.sleep(5)
                    if page.is_closed():
                        raise RuntimeError("抖音窗口被关闭，采集已停止")
                    if "douyin.com/search" not in page.url:
                        _safe_goto(page, search_url)
                    collect_deadline = time.time() + COLLECT_WAIT_SECONDS
                    while time.time() < collect_deadline:
                        videos = _extract_videos(api_responses, keyword, max_count)
                        if videos:
                            break
                        time.sleep(2)
                    if videos:
                        break
                ctx.close()

            if not videos:
                raise RuntimeError("未采集到搜索结果。请确认抖音窗口已登录，并能正常看到视频搜索列表。")

            out_dir = os.path.join(SETTINGS.output_dir, "素材导入", time.strftime("%Y-%m-%d"), f"douyin_search_{safe_kw}")
            os.makedirs(out_dir, exist_ok=True)
            archived = []
            for i, video in enumerate(videos):
                safe_desc = _safe_filename(str(video.get("desc", "")), "untitled")[:36].strip(" ._")
                vdir = _unique_dir(out_dir, f"{i + 1:02d}_{safe_desc}")
                os.makedirs(vdir, exist_ok=True)
                mp4_path = os.path.join(vdir, f"{i + 1:02d}.mp4")
                cover_path = os.path.join(vdir, f"{i + 1:02d}_cover.jpg")
                _download_if_possible(video.get("video_url", ""), mp4_path, 30)
                _download_if_possible(video.get("cover_url", ""), cover_path, 15)
                video = dict(video)
                video["archive_dir"] = vdir
                video["file_path"] = mp4_path if os.path.exists(mp4_path) else ""
                video["cover_path"] = cover_path if os.path.exists(cover_path) else ""
                with open(os.path.join(vdir, "metadata.json"), "w", encoding="utf-8") as f:
                    json.dump(video, f, ensure_ascii=False, indent=2)
                archived.append(video)

            with open(res_path, "w", encoding="utf-8") as f:
                json.dump({"ok": True, "count": len(archived), "videos": archived, "out_dir": out_dir}, f, ensure_ascii=False)
            _write_status(stat_path, "done")
        except Exception as exc:
            _write_status(stat_path, f"error:{exc}")

    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "job_id": job_id, "message": f"开始搜索「{keyword}」。如果抖音要求登录，请先在弹出的浏览器里登录，系统会继续采集。"}


def import_status(job_id: str) -> dict[str, object]:
    stat_path, res_path = _job_paths(job_id)
    if not os.path.exists(stat_path):
        return {"ok": True, "status": "pending"}
    with open(stat_path, encoding="utf-8") as f:
        status = f.read().strip()
    result = None
    if status == "done" and os.path.exists(res_path):
        with open(res_path, encoding="utf-8") as f:
            result = json.load(f)
    return {"ok": True, "status": status, "result": result}


def record_to_form(record: dict[str, object]) -> dict[str, str]:
    metrics = [
        f"播放: {record.get('play_count') or record.get('view_count') or ''}",
        f"点赞: {record.get('like_count') or ''}",
        f"评论: {record.get('comment_count') or ''}",
        f"收藏: {record.get('collect_count') or ''}",
        f"转发: {record.get('share_count') or ''}",
        f"时长: {record.get('duration') or ''}秒",
    ]
    return {
        "title": str(record.get("title") or record.get("desc") or ""),
        "platform": str(record.get("platform") or ""),
        "url": str(record.get("page_url") or record.get("video_url") or ""),
        "metrics": "\n".join(x for x in metrics if not x.endswith(": ") and not x.endswith(": 秒")),
        "sales": str(record.get("sales") or ""),
        "file_path": str(record.get("file_path") or ""),
        "notes": f"作者: {record.get('author_name') or ''}\n归档: {record.get('archive_dir') or ''}".strip(),
    }
